import os
import sys
import json
import uuid
import queue
import hashlib
import threading
import logging
import traceback
from logging.handlers import RotatingFileHandler
from flask import (Flask, render_template, request, jsonify, g,
                   Response, stream_with_context, abort, send_file, redirect, url_for)
from werkzeug.utils import secure_filename
from werkzeug.exceptions import RequestEntityTooLarge
from sqlalchemy import or_, text
from models import db, Project, PdfDocument, Requirement, TOCItem, BusinessInfo, ReqTypeRule, ProjectAttachment, ATTACHMENT_SLOTS, ProjectFileAttachment, PROJECT_FILE_CATEGORIES, ProposalAnalysis, RequirementProposalImage, toc_requirement, TodoItem, WorkLog, VectorChunk
from gemini_service import extract_all
from proposal_image_service import ProposalImageDependencyError, render_requirement_proposal_image
from dotenv import load_dotenv

load_dotenv()

# ── 경로 설정 (일반 실행 / PyInstaller 실행 파일 양쪽 지원) ────────────────────
_FROZEN = getattr(sys, 'frozen', False)
if _FROZEN:
    _BASE_DIR = os.path.dirname(sys.executable)
    _TEMPLATE_DIR = os.path.join(sys._MEIPASS, 'templates')
else:
    _BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    _TEMPLATE_DIR = os.path.join(_BASE_DIR, 'templates')

_INSTANCE_DIR = os.path.join(_BASE_DIR, 'instance')
os.makedirs(_INSTANCE_DIR, exist_ok=True)

app = Flask(__name__, template_folder=_TEMPLATE_DIR)

# ── 로깅 설정 ──────────────────────────────────────────────────────────────────
LOG_DIR = os.path.join(_BASE_DIR, 'logs')
os.makedirs(LOG_DIR, exist_ok=True)

# 재시작 시 로그 파일 초기화
_log_path = os.path.join(LOG_DIR, 'app.log')
open(_log_path, 'w', encoding='utf-8').close()

_file_handler = RotatingFileHandler(
    _log_path,
    maxBytes=5 * 1024 * 1024,  # 5MB
    backupCount=3,
    encoding='utf-8',
)
_file_handler.setLevel(logging.INFO)
_file_handler.setFormatter(logging.Formatter(
    '[%(asctime)s] %(levelname)s %(module)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
))

_console_handler = logging.StreamHandler()
_console_handler.setLevel(logging.DEBUG)
_console_handler.setFormatter(logging.Formatter(
    '[%(asctime)s] %(levelname)s: %(message)s', datefmt='%H:%M:%S'
))

logging.basicConfig(level=logging.DEBUG, handlers=[_file_handler, _console_handler])
logger = logging.getLogger(__name__)

# werkzeug 로그도 파일에 기록
logging.getLogger('werkzeug').addHandler(_file_handler)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret')
_default_db = f'sqlite:///{os.path.join(_INSTANCE_DIR, "req_manager.db")}'
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', _default_db)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,   # 끊어진 연결 자동 재연결
    'pool_recycle': 3600,    # 1시간마다 연결 갱신
}
app.config['UPLOAD_FOLDER'] = os.path.join(_BASE_DIR, 'uploads')
UPLOAD_LIMIT_MB = int(os.environ.get('MAX_UPLOAD_MB', '300'))
app.config['MAX_CONTENT_LENGTH'] = UPLOAD_LIMIT_MB * 1024 * 1024

GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY', '')
SAVED_DIR = os.path.join(_BASE_DIR, 'uploads', 'saved')

# ── 로컬 인증 (MAC 주소 기반) ──────────────────────────────────────────────────
def _gen_local_token() -> str:
    """이 머신의 MAC 주소 + SECRET_KEY 로 고유 토큰 생성"""
    mac_int = uuid.getnode()
    mac_str = ':'.join(f'{(mac_int >> (i * 8)) & 0xff:02x}' for i in range(5, -1, -1))
    secret  = app.config.get('SECRET_KEY', 'dev-secret')
    return hashlib.sha256(f"{mac_str}:{secret}".encode()).hexdigest()[:40]

_LOCAL_TOKEN    = _gen_local_token()
_ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', '')
_PROTECTED_PREFIXES = ('/todos', '/worklog', '/chat', '/calendar',
                        '/api/todos', '/api/worklogs', '/api/chat', '/api/dashboard')

def _collect_local_ips() -> set:
    """이 머신에 할당된 모든 IP 주소 수집 (127.0.0.1, LAN IP 등)"""
    import socket
    ips = {'127.0.0.1', '::1'}
    try:
        hostname = socket.gethostname()
        for info in socket.getaddrinfo(hostname, None):
            ips.add(info[4][0])
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
            s.connect(('8.8.8.8', 80))
            ips.add(s.getsockname()[0])
    except Exception:
        pass
    return ips

_LOCAL_IPS = _collect_local_ips()
logger.info("로컬 허용 IP 목록: %s", _LOCAL_IPS)

def _is_local_request() -> bool:
    return request.remote_addr in _LOCAL_IPS

def _has_local_auth() -> bool:
    return request.cookies.get('_lat') == _LOCAL_TOKEN

def _is_admin() -> bool:
    """로컬 IP, 로컬 쿠키, 또는 세션 로그인 중 하나면 관리자"""
    from flask import session as _sess
    return _is_local_request() or _has_local_auth() or _sess.get('is_admin') is True

@app.before_request
def _check_local_auth():
    """보호된 경로 접근 시 인증 확인"""
    if request.path in ('/login', '/logout'):
        return
    if not any(request.path.startswith(p) for p in _PROTECTED_PREFIXES):
        return
    if _is_admin():
        g.set_local_cookie = _is_local_request() and not _has_local_auth()
        return
    if request.path.startswith('/api/'):
        return jsonify({'error': '접근 권한이 없습니다.'}), 403
    return redirect(f'/login?next={request.path}')

@app.after_request
def _issue_local_cookie(response):
    """로컬호스트 첫 접근 시 인증 쿠키 자동 발급"""
    if getattr(g, 'set_local_cookie', False):
        response.set_cookie(
            '_lat', _LOCAL_TOKEN,
            httponly=True, samesite='Strict',
            max_age=365 * 24 * 3600,
        )
    return response

@app.context_processor
def _inject_local_auth():
    """템플릿에서 is_local_user / is_session_login 변수 사용 가능"""
    from flask import session as _sess
    return {
        'is_local_user':    _is_admin(),
        'is_session_login': _sess.get('is_admin') is True and not _is_local_request(),
    }

# ── 로그인 / 로그아웃 ─────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login_page():
    from flask import session as _sess
    if _is_admin():
        return redirect('/')
    error = None
    if request.method == 'POST':
        pw = request.form.get('password', '')
        if _ADMIN_PASSWORD and pw == _ADMIN_PASSWORD:
            _sess['is_admin'] = True
            _sess.permanent = True
            next_url = request.args.get('next', '/')
            return redirect(next_url)
        error = '비밀번호가 올바르지 않습니다.'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout_page():
    from flask import session as _sess
    _sess.pop('is_admin', None)
    return redirect('/')
ATTACH_DIR = os.path.join(_BASE_DIR, 'uploads', 'attachments')
GENERATED_DIR = os.path.join(_BASE_DIR, 'uploads', 'generated')
PROJECT_FILE_DIR = os.path.join(_BASE_DIR, 'uploads', 'project_files')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(SAVED_DIR, exist_ok=True)
os.makedirs(ATTACH_DIR, exist_ok=True)
os.makedirs(GENERATED_DIR, exist_ok=True)
os.makedirs(PROJECT_FILE_DIR, exist_ok=True)

db.init_app(app)


def _migrate(conn):
    """기존 DB에 누락된 컬럼/테이블을 자동으로 추가."""
    migrations = [
        "ALTER TABLE pdf_document ADD COLUMN saved_filename VARCHAR(500)",
        """CREATE TABLE IF NOT EXISTS proposal_analysis (
            id INTEGER PRIMARY KEY,
            created_at DATETIME,
            rfp_name VARCHAR(500) NOT NULL,
            proposal_name VARCHAR(500) NOT NULL,
            results_json TEXT NOT NULL,
            summary_json TEXT,
            total_count INTEGER DEFAULT 0,
            full_count INTEGER DEFAULT 0,
            partial_count INTEGER DEFAULT 0,
            missing_count INTEGER DEFAULT 0
        )""",
        """CREATE TABLE IF NOT EXISTS requirement_proposal_image (
            id INTEGER PRIMARY KEY,
            requirement_id INTEGER NOT NULL REFERENCES requirement(id) ON DELETE CASCADE,
            orientation VARCHAR(20) DEFAULT 'landscape',
            template_type VARCHAR(50) DEFAULT 'auto',
            tone VARCHAR(50) DEFAULT 'public',
            title VARCHAR(500),
            saved_filename VARCHAR(500) NOT NULL,
            payload_json TEXT,
            created_at DATETIME
        )""",
        """CREATE TABLE IF NOT EXISTS req_type_rule (
            id INTEGER PRIMARY KEY,
            prefix VARCHAR(50) NOT NULL,
            label VARCHAR(100) NOT NULL,
            bg_color VARCHAR(20) DEFAULT '#dbeafe',
            text_color VARCHAR(20) DEFAULT '#1d4ed8',
            order_index INTEGER DEFAULT 0
        )""",
        """CREATE TABLE IF NOT EXISTS business_info (
            id INTEGER PRIMARY KEY,
            pdf_id INTEGER NOT NULL UNIQUE REFERENCES pdf_document(id) ON DELETE CASCADE,
            business_name VARCHAR(500),
            business_cost VARCHAR(200),
            business_period VARCHAR(200),
            client VARCHAR(500),
            contractor VARCHAR(500),
            overview TEXT,
            extras TEXT
        )""",
        """CREATE TABLE IF NOT EXISTS todo_item (
            id INTEGER PRIMARY KEY,
            title VARCHAR(500) NOT NULL,
            description TEXT,
            priority VARCHAR(20) DEFAULT '보통',
            status VARCHAR(20) DEFAULT '할일',
            due_date VARCHAR(20),
            order_index INTEGER DEFAULT 0,
            created_at DATETIME,
            updated_at DATETIME
        )""",
        """CREATE TABLE IF NOT EXISTS work_log (
            id INTEGER PRIMARY KEY,
            title VARCHAR(500) NOT NULL,
            content TEXT,
            tags TEXT,
            created_at DATETIME,
            updated_at DATETIME
        )""",
        """CREATE TABLE IF NOT EXISTS vector_chunk (
            id INTEGER PRIMARY KEY,
            source_type VARCHAR(50) NOT NULL,
            source_id INTEGER NOT NULL,
            chunk_text TEXT NOT NULL,
            source_url VARCHAR(500),
            source_label VARCHAR(500),
            embedding TEXT,
            updated_at DATETIME
        )""",
        """CREATE TABLE IF NOT EXISTS project_file_attachment (
            id INTEGER PRIMARY KEY,
            project_id INTEGER NOT NULL REFERENCES project(id) ON DELETE CASCADE,
            category VARCHAR(30) NOT NULL,
            original_name VARCHAR(500) NOT NULL,
            saved_filename VARCHAR(500) NOT NULL,
            uploaded_at DATETIME
        )""",
        # unique constraint 추가 (없으면 생성)
        "CREATE UNIQUE INDEX IF NOT EXISTS uq_vector_chunk_src ON vector_chunk(source_type, source_id)",
        "ALTER TABLE vector_chunk ADD COLUMN source_date DATETIME",
        "ALTER TABLE todo_item ADD COLUMN start_date VARCHAR(20)",
        "ALTER TABLE requirement ADD COLUMN status VARCHAR(20) DEFAULT '신규'",
        "ALTER TABLE requirement ADD COLUMN priority VARCHAR(20) DEFAULT '보통'",
        "ALTER TABLE requirement ADD COLUMN note TEXT",
    ]
    for sql in migrations:
        try:
            conn.execute(text(sql))
        except Exception:
            pass  # 이미 존재하면 무시

    # vector_chunk 중복 제거 (unique index 생성 실패한 경우 먼저 정리)
    try:
        conn.execute(text("""
            DELETE FROM vector_chunk WHERE id NOT IN (
                SELECT MIN(id) FROM vector_chunk GROUP BY source_type, source_id
            )
        """))
        conn.execute(text(
            "CREATE UNIQUE INDEX IF NOT EXISTS uq_vector_chunk_src ON vector_chunk(source_type, source_id)"
        ))
    except Exception:
        pass


DEFAULT_REQ_TYPES = [
    ('SFR', '기능요구사항',        '#dbeafe', '#1d4ed8'),
    ('FRQ', '기능요구사항',        '#dbeafe', '#1d4ed8'),
    ('FR',  '기능요구사항',        '#dbeafe', '#1d4ed8'),
    ('SQR', '품질요구사항',        '#dcfce7', '#166534'),
    ('QR',  '품질요구사항',        '#dcfce7', '#166534'),
    ('NFR', '품질요구사항',        '#dcfce7', '#166534'),
    ('SCR', '제약사항',            '#fef3c7', '#92400e'),
    ('CR',  '제약사항',            '#fef3c7', '#92400e'),
    ('DAR', '데이터요구사항',      '#f3e8ff', '#6b21a8'),
    ('DR',  '데이터요구사항',      '#f3e8ff', '#6b21a8'),
    ('PMR', '성능요구사항',        '#fce7f3', '#9d174d'),
    ('PR',  '성능요구사항',        '#fce7f3', '#9d174d'),
    ('IFR', '인터페이스요구사항',  '#e0f2fe', '#075985'),
    ('IF',  '인터페이스요구사항',  '#e0f2fe', '#075985'),
    ('SER', '보안요구사항',        '#fff7ed', '#9a3412'),
    ('SEC', '보안요구사항',        '#fff7ed', '#9a3412'),
    ('TCR', '기술요구사항',        '#f0fdf4', '#15803d'),
    ('TR',  '기술요구사항',        '#f0fdf4', '#15803d'),
    ('REQ', '요구사항',            '#f1f5f9', '#475569'),
    ('RQ',  '요구사항',            '#f1f5f9', '#475569'),
]

def _set_low_priority():
    """현재 스레드를 낮은 우선순위로 설정 (Windows)"""
    try:
        import ctypes
        # THREAD_PRIORITY_LOWEST = -2
        ctypes.windll.kernel32.SetThreadPriority(
            ctypes.windll.kernel32.GetCurrentThread(), -2
        )
    except Exception:
        pass


def _startup_sync():
    """앱 시작 시 모델 프리웜 + 누락된 임베딩 자동 보완 (백그라운드, 낮은 우선순위)"""
    import time as _time
    _set_low_priority()          # 시스템 영향 최소화
    _time.sleep(30)              # 로그인 직후 다른 프로그램 안정화 대기

    # 1) 임베딩 모델 미리 로드 (첫 채팅 시 CPU 스파이크 방지)
    try:
        from rag_service import _get_model
        _get_model()
        logger.info("RAG: 임베딩 모델 프리웜 완료")
    except Exception as e:
        logger.warning("RAG: 모델 프리웜 실패: %s", e)

    # 2) 누락된 임베딩 보완
    with app.app_context():
        try:
            from rag_service import build_chunks, embed_document
            from datetime import datetime as _dt
            chunks     = build_chunks()
            synced_ids = {
                (c.source_type, c.source_id)
                for c in VectorChunk.query.with_entities(
                    VectorChunk.source_type, VectorChunk.source_id).all()
            }
            missing = [c for c in chunks
                       if (c['source_type'], c['source_id']) not in synced_ids]
            if not missing:
                logger.info("RAG: 누락 항목 없음 (%d개 동기화됨)", len(synced_ids))
                return
            logger.info("RAG: 누락 항목 %d개 자동 동기화 시작", len(missing))
            for c in missing:
                try:
                    emb = embed_document(c['chunk_text'])
                    vc  = VectorChunk(
                        source_type  = c['source_type'],
                        source_id    = c['source_id'],
                        chunk_text   = c['chunk_text'],
                        source_url   = c['source_url'],
                        source_label = c['source_label'],
                        embedding    = json.dumps(emb),
                        updated_at   = _dt.utcnow(),
                        source_date  = c.get('source_date'),
                    )
                    db.session.add(vc)
                    db.session.commit()
                    _time.sleep(0.05)  # 청크당 50ms 대기 — CPU 과점 방지
                except Exception as e:
                    logger.warning("시작 시 청크 동기화 실패: %s", e)
            logger.info("RAG: 시작 시 자동 동기화 완료 (%d개)", len(missing))
        except Exception as e:
            logger.warning("RAG 시작 시 동기화 오류: %s", e)


with app.app_context():
    db.create_all()
    with db.engine.connect() as conn:
        _migrate(conn)
        conn.commit()
    if ReqTypeRule.query.count() == 0:
        for i, (prefix, label, bg, fg) in enumerate(DEFAULT_REQ_TYPES):
            db.session.add(ReqTypeRule(prefix=prefix, label=label, bg_color=bg, text_color=fg, order_index=i))
        db.session.commit()
    logger.info("DB 초기화 완료")
    # 앱 시작 시 누락된 임베딩 백그라운드 동기화
    # Flask debug+reloader 환경에서 이중 실행 방지:
    #   - _FROZEN=True (운영 exe): reloader 없이 단일 프로세스 → 항상 실행
    #   - WERKZEUG_RUN_MAIN='true': reloader 자식 프로세스(실제 앱) → 실행
    #   - WERKZEUG_RUN_MAIN=None + _FROZEN=False: reloader 부모 프로세스 → 건너뜀
    if _FROZEN or os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
        threading.Thread(target=_startup_sync, daemon=True).start()


# ── 글로벌 에러 핸들러 ─────────────────────────────────────────────────────────

@app.errorhandler(RequestEntityTooLarge)
def handle_request_entity_too_large(e):
    return jsonify({
        'error': f'파일 크기가 너무 큽니다. 최대 {UPLOAD_LIMIT_MB}MB까지 업로드할 수 있습니다.'
    }), 413


@app.errorhandler(Exception)
def handle_exception(e):
    tb = traceback.format_exc()
    logger.error("Unhandled exception on %s %s\n%s", request.method, request.path, tb)
    if app.debug:
        return jsonify({'error': str(e), 'traceback': tb}), 500
    return jsonify({'error': '서버 오류가 발생했습니다. logs/app.log를 확인하세요.'}), 500


@app.errorhandler(404)
def handle_404(e):
    if request.path.startswith('/api/'):
        return jsonify({'error': '리소스를 찾을 수 없습니다.'}), 404
    return render_template('404.html'), 404


_task_queues: dict[str, queue.Queue] = {}


# ── RAG 자동 동기화 헬퍼 ──────────────────────────────────────────────────────

def _upsert_chunk(source_type: str, source_id: int):
    """특정 항목의 벡터 청크를 백그라운드에서 생성/갱신"""

    def run():
        _set_low_priority()
        with app.app_context():
            from rag_service import build_chunk_for, embed_document
            from datetime import datetime as _dt
            try:
                data = build_chunk_for(source_type, source_id)
                if not data:
                    VectorChunk.query.filter_by(
                        source_type=source_type, source_id=source_id).delete()
                    db.session.commit()
                    return
                emb = embed_document(data['chunk_text'])
                vc  = VectorChunk.query.filter_by(
                    source_type=source_type, source_id=source_id).first()
                if not vc:
                    vc = VectorChunk(source_type=source_type, source_id=source_id)
                    db.session.add(vc)
                vc.chunk_text   = data['chunk_text']
                vc.source_url   = data['source_url']
                vc.source_label = data['source_label']
                vc.embedding    = json.dumps(emb)
                vc.updated_at   = _dt.utcnow()
                vc.source_date  = data.get('source_date')
                db.session.commit()
                logger.debug("청크 자동 갱신: %s:%s", source_type, source_id)
            except Exception as e:
                logger.warning("청크 자동 갱신 실패 [%s:%s]: %s", source_type, source_id, e)

    threading.Thread(target=run, daemon=True).start()


def _delete_chunk(source_type: str, source_id: int):
    """항목 삭제 시 벡터 청크도 삭제"""
    def run():
        with app.app_context():
            try:
                VectorChunk.query.filter_by(
                    source_type=source_type, source_id=source_id).delete()
                db.session.commit()
            except Exception as e:
                logger.warning("청크 삭제 실패 [%s:%s]: %s", source_type, source_id, e)
    threading.Thread(target=run, daemon=True).start()


_analyze_results: dict[str, list] = {}   # task_id -> 매칭 결과
_verify_results:  dict[str, list] = {}   # task_id -> 검증 결과


# ── Background worker ─────────────────────────────────────────────────────────

def _process_pdf(task_id: str, temp_path: str, original_name: str, project_id: int):
    q = _task_queues[task_id]
    logger.info("[task:%s] PDF 처리 시작: %s (project_id=%s)", task_id[:8], original_name, project_id)

    def send(step: int, msg: str):
        logger.info("[task:%s] step %d: %s", task_id[:8], step, msg)
        q.put({'type': 'progress', 'step': step, 'total': 7, 'msg': msg})

    saved_filename = f"{uuid.uuid4().hex}.pdf"
    saved_path = os.path.join(SAVED_DIR, saved_filename)

    try:
        requirements_data, toc_data, business_info = extract_all(
            temp_path, GEMINI_API_KEY,
            progress_cb=lambda step, msg: send(step, msg),
        )
        logger.info("[task:%s] Gemini 추출 완료: req=%d, toc=%d",
                    task_id[:8], len(requirements_data), len(toc_data))

        send(6, f'데이터베이스에 저장 중... (요구사항 {len(requirements_data)}개, 목차 {len(toc_data)}개)')

        # Move temp file to permanent storage
        os.rename(temp_path, saved_path)

        with app.app_context():
            pdf_doc = PdfDocument(
                project_id=project_id,
                original_name=original_name,
                saved_filename=saved_filename,
            )
            db.session.add(pdf_doc)
            db.session.flush()

            # Save business info
            if isinstance(business_info, dict):
                extras = business_info.get('extras') or {}
                db.session.add(BusinessInfo(
                    pdf_id=pdf_doc.id,
                    business_name=str(business_info.get('business_name') or '').strip() or None,
                    business_cost=str(business_info.get('business_cost') or '').strip() or None,
                    business_period=str(business_info.get('business_period') or '').strip() or None,
                    client=str(business_info.get('client') or '').strip() or None,
                    contractor=str(business_info.get('contractor') or '').strip() or None,
                    overview=str(business_info.get('overview') or '').strip() or None,
                    extras=json.dumps(extras, ensure_ascii=False) if extras else None,
                ))

            req_count = 0
            for item in requirements_data:
                if not isinstance(item, dict):
                    continue
                req_id = str(item.get('req_id', '')).strip()
                if not req_id:
                    continue
                db.session.add(Requirement(
                    pdf_id=pdf_doc.id,
                    req_id=req_id,
                    req_name=str(item.get('req_name', '')).strip(),
                    detail=str(item.get('detail', '')).strip(),
                ))
                req_count += 1
            db.session.flush()

            toc_count = 0
            for i, item in enumerate(toc_data):
                if not isinstance(item, dict):
                    continue
                d1 = (item.get('depth1') or '').strip() or None
                d2 = (item.get('depth2') or '').strip() or None
                d3 = (item.get('depth3') or '').strip() or None
                if not d1 and not d2 and not d3:
                    continue
                db.session.add(TOCItem(
                    pdf_id=pdf_doc.id,
                    depth1=d1, depth2=d2, depth3=d3,
                    page_number=str(item.get('page_number') or '').strip() or None,
                    work_status='신규',
                    order_index=i,
                ))
                toc_count += 1

            db.session.commit()
            pdf_id = pdf_doc.id

        send(7, '완료!')
        logger.info("[task:%s] 처리 완료: pdf_id=%s, req=%d, toc=%d",
                    task_id[:8], pdf_id, req_count, toc_count)
        q.put({
            'type': 'done',
            'msg': f'완료 — 요구사항 {req_count}개, 목차 {toc_count}개',
            'req_count': req_count,
            'toc_count': toc_count,
            'pdf_id': pdf_id,
        })

    except Exception as e:
        tb = traceback.format_exc()
        logger.error("[task:%s] PDF 처리 실패: %s\n%s", task_id[:8], original_name, tb)
        with app.app_context():
            db.session.rollback()
        if os.path.exists(saved_path):
            os.remove(saved_path)
        q.put({'type': 'error', 'msg': str(e)})
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


# ── Pages ─────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    projects = Project.query.order_by(Project.created_at.desc()).all()
    return render_template('index.html', projects=projects)


@app.route('/project/<int:project_id>')
def project_detail(project_id):
    project = db.get_or_404(Project, project_id)
    # ?manage=1 이면 관리 화면 (PDF 추가, 첨부파일 등)
    if project.pdfs and not request.args.get('manage'):
        latest = max(project.pdfs, key=lambda p: p.created_at or 0)
        return redirect(url_for('pdf_detail', project_id=project_id, pdf_id=latest.id))
    return render_template('project.html', project=project)


@app.route('/project/<int:project_id>/pdf/<int:pdf_id>')
def pdf_detail(project_id, pdf_id):
    project = db.get_or_404(Project, project_id)
    pdf = db.get_or_404(PdfDocument, pdf_id)
    if pdf.project_id != project_id:
        abort(404)
    return render_template('pdf_detail.html', project=project, pdf=pdf)


@app.route('/search')
def search_page():
    return render_template('search.html')


# ── Project API ───────────────────────────────────────────────────────────────

@app.route('/api/projects', methods=['POST'])
def api_project_create():
    data = request.get_json()
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': '프로젝트 이름을 입력해주세요.'}), 400
    p = Project(name=name, description=(data.get('description') or '').strip() or None)
    db.session.add(p)
    db.session.commit()
    return jsonify({'success': True, 'project': p.to_dict()})


@app.route('/api/projects/<int:pid>', methods=['PUT'])
def api_project_update(pid):
    p = db.get_or_404(Project, pid)
    data = request.get_json()
    p.name = (data.get('name') or p.name).strip()
    p.description = (data.get('description') or '').strip() or None
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/projects/<int:pid>', methods=['DELETE'])
def api_project_delete(pid):
    p = db.get_or_404(Project, pid)
    # Remove saved PDF files
    for pdf in p.pdfs:
        if pdf.saved_filename:
            fp = os.path.join(SAVED_DIR, pdf.saved_filename)
            if os.path.exists(fp):
                os.remove(fp)
    for att in p.attachments:
        fp = os.path.join(ATTACH_DIR, att.saved_filename)
        if os.path.exists(fp):
            os.remove(fp)
    for item in p.file_attachments:
        fp = os.path.join(PROJECT_FILE_DIR, item.saved_filename)
        if os.path.exists(fp):
            os.remove(fp)
    db.session.delete(p)
    db.session.commit()
    return jsonify({'success': True})


# ── Project Attachment API ────────────────────────────────────────────────────

@app.route('/api/project/<int:project_id>/attachments')
def api_attachments_list(project_id):
    db.get_or_404(Project, project_id)
    rows = ProjectAttachment.query.filter_by(project_id=project_id).all()
    by_slot = {r.slot: r.to_dict() for r in rows}
    result = {}
    for slot, (cat, kind) in ATTACHMENT_SLOTS.items():
        result[slot] = by_slot.get(slot)
    return jsonify(result)


@app.route('/api/project/<int:project_id>/attachment/<slot>', methods=['POST'])
def api_attachment_upload(project_id, slot):
    if slot not in ATTACHMENT_SLOTS:
        abort(400, 'Invalid slot')
    db.get_or_404(Project, project_id)

    f = request.files.get('file')
    if not f or not f.filename:
        abort(400, 'No file')

    # 기존 파일 삭제
    existing = ProjectAttachment.query.filter_by(project_id=project_id, slot=slot).first()
    if existing:
        old_path = os.path.join(ATTACH_DIR, existing.saved_filename)
        if os.path.exists(old_path):
            os.remove(old_path)
        db.session.delete(existing)
        db.session.flush()

    ext = os.path.splitext(secure_filename(f.filename))[1]
    saved_name = f'{uuid.uuid4().hex}{ext}'
    f.save(os.path.join(ATTACH_DIR, saved_name))

    att = ProjectAttachment(
        project_id=project_id,
        slot=slot,
        original_name=f.filename,
        saved_filename=saved_name,
    )
    db.session.add(att)
    db.session.commit()
    return jsonify(att.to_dict())


@app.route('/api/project/<int:project_id>/attachment/<slot>/file')
def api_attachment_file(project_id, slot):
    att = ProjectAttachment.query.filter_by(project_id=project_id, slot=slot).first_or_404()
    path = os.path.join(ATTACH_DIR, att.saved_filename)
    if not os.path.exists(path):
        abort(404)
    return send_file(path, download_name=att.original_name, as_attachment=True)


@app.route('/api/project/<int:project_id>/attachment/<slot>', methods=['DELETE'])
def api_attachment_delete(project_id, slot):
    att = ProjectAttachment.query.filter_by(project_id=project_id, slot=slot).first_or_404()
    path = os.path.join(ATTACH_DIR, att.saved_filename)
    if os.path.exists(path):
        try:
            os.remove(path)
        except PermissionError:
            return jsonify({'error': '파일이 사용 중이라 삭제할 수 없습니다. 파일을 닫은 뒤 다시 시도하세요.'}), 409
    db.session.delete(att)
    db.session.commit()
    return jsonify({'success': True})


# ── Project File Box API ─────────────────────────────────────────────────────

@app.route('/api/project/<int:project_id>/filebox')
def api_project_filebox_list(project_id):
    db.get_or_404(Project, project_id)
    rows = ProjectFileAttachment.query.filter_by(project_id=project_id) \
        .order_by(ProjectFileAttachment.uploaded_at.desc(), ProjectFileAttachment.id.desc()).all()
    result = {
        key: {
            'label': label,
            'description': description,
            'files': [],
        }
        for key, (label, description) in PROJECT_FILE_CATEGORIES.items()
    }
    for row in rows:
        if row.category not in result:
            result[row.category] = {
                'label': row.category,
                'description': '',
                'files': [],
            }
        result[row.category]['files'].append(row.to_dict())
    return jsonify(result)


@app.route('/api/project/<int:project_id>/filebox/<category>', methods=['POST'])
def api_project_filebox_upload(project_id, category):
    if category not in PROJECT_FILE_CATEGORIES:
        return jsonify({'error': '잘못된 파일함 영역입니다.'}), 400
    db.get_or_404(Project, project_id)

    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'error': '파일을 선택하세요.'}), 400

    ext = os.path.splitext(secure_filename(f.filename))[1]
    saved_name = f'{uuid.uuid4().hex}{ext}'
    f.save(os.path.join(PROJECT_FILE_DIR, saved_name))

    item = ProjectFileAttachment(
        project_id=project_id,
        category=category,
        original_name=f.filename,
        saved_filename=saved_name,
    )
    db.session.add(item)
    db.session.commit()
    return jsonify(item.to_dict()), 201


@app.route('/api/project/<int:project_id>/filebox/<int:file_id>/file')
def api_project_filebox_file(project_id, file_id):
    item = ProjectFileAttachment.query.filter_by(
        id=file_id, project_id=project_id).first_or_404()
    path = os.path.join(PROJECT_FILE_DIR, item.saved_filename)
    if not os.path.exists(path):
        abort(404)
    return send_file(path, download_name=item.original_name, as_attachment=True)


@app.route('/api/project/<int:project_id>/filebox/<int:file_id>', methods=['DELETE'])
def api_project_filebox_delete(project_id, file_id):
    item = ProjectFileAttachment.query.filter_by(
        id=file_id, project_id=project_id).first_or_404()
    path = os.path.join(PROJECT_FILE_DIR, item.saved_filename)
    if os.path.exists(path):
        try:
            os.remove(path)
        except PermissionError:
            return jsonify({'error': '파일이 사용 중이라 삭제할 수 없습니다. 파일을 닫은 뒤 다시 시도하세요.'}), 409
    db.session.delete(item)
    db.session.commit()
    return jsonify({'success': True})


# ── Upload & SSE ──────────────────────────────────────────────────────────────

@app.route('/project/<int:project_id>/upload', methods=['POST'])
def upload(project_id):
    db.get_or_404(Project, project_id)
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다.'}), 400
    file = request.files['file']
    if not file.filename or not file.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'PDF 파일만 업로드 가능합니다.'}), 400

    original_name = file.filename
    temp_filename = f"tmp_{uuid.uuid4().hex}_{secure_filename(original_name)}"
    temp_path = os.path.join(app.config['UPLOAD_FOLDER'], temp_filename)
    file.save(temp_path)

    task_id = str(uuid.uuid4())
    _task_queues[task_id] = queue.Queue()
    threading.Thread(
        target=_process_pdf,
        args=(task_id, temp_path, original_name, project_id),
        daemon=True,
    ).start()
    return jsonify({'task_id': task_id})


@app.route('/upload/progress/<task_id>')
def upload_progress(task_id):
    def generate():
        q = _task_queues.get(task_id)
        if not q:
            yield f"data: {json.dumps({'type': 'error', 'msg': '작업을 찾을 수 없습니다.'})}\n\n"
            return
        try:
            while True:
                try:
                    msg = q.get(timeout=180)
                    yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"
                    if msg.get('type') in ('done', 'error'):
                        break
                except queue.Empty:
                    yield 'data: {"type":"keepalive"}\n\n'
        finally:
            _task_queues.pop(task_id, None)

    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'},
    )


# ── PDF File Serving ──────────────────────────────────────────────────────────

@app.route('/api/pdf/<int:pdf_id>/file')
def api_pdf_file(pdf_id):
    pdf = db.get_or_404(PdfDocument, pdf_id)
    if not pdf.saved_filename:
        abort(404)
    filepath = os.path.join(SAVED_DIR, pdf.saved_filename)
    if not os.path.exists(filepath):
        abort(404)
    return send_file(filepath, mimetype='application/pdf',
                     download_name=pdf.original_name)


@app.route('/api/pdf/<int:pdf_id>', methods=['DELETE'])
def api_pdf_delete(pdf_id):
    pdf = db.get_or_404(PdfDocument, pdf_id)
    if pdf.saved_filename:
        fp = os.path.join(SAVED_DIR, pdf.saved_filename)
        if os.path.exists(fp):
            os.remove(fp)
    db.session.delete(pdf)
    db.session.commit()
    return jsonify({'success': True})


# ── Business Info API ────────────────────────────────────────────────────────

@app.route('/api/pdf/<int:pdf_id>/business_info')
def api_business_info(pdf_id):
    pdf = db.get_or_404(PdfDocument, pdf_id)
    if not pdf.business_info:
        return jsonify(None)
    return jsonify(pdf.business_info.to_dict())


@app.route('/api/pdf/<int:pdf_id>/business_info', methods=['PUT'])
def api_business_info_update(pdf_id):
    pdf = db.get_or_404(PdfDocument, pdf_id)
    data = request.get_json()
    bi = pdf.business_info
    if not bi:
        bi = BusinessInfo(pdf_id=pdf_id)
        db.session.add(bi)
    bi.business_name = data.get('business_name') or None
    bi.business_cost = data.get('business_cost') or None
    bi.business_period = data.get('business_period') or None
    bi.client = data.get('client') or None
    bi.contractor = data.get('contractor') or None
    bi.overview = data.get('overview') or None
    extras = data.get('extras')
    bi.extras = json.dumps(extras, ensure_ascii=False) if extras else None
    db.session.commit()
    return jsonify({'success': True})


# ── Requirements API ──────────────────────────────────────────────────────────

@app.route('/api/pdf/<int:pdf_id>/requirements')
def api_requirements(pdf_id):
    db.get_or_404(PdfDocument, pdf_id)
    items = Requirement.query.filter_by(pdf_id=pdf_id).order_by(Requirement.id).all()
    return jsonify([r.to_dict() for r in items])


@app.route('/api/requirements/<int:rid>', methods=['PUT'])
def api_req_update(rid):
    r = db.get_or_404(Requirement, rid)
    data = request.get_json()
    r.req_id = (data.get('req_id') or r.req_id).strip()
    r.req_name = (data.get('req_name') or r.req_name).strip()
    r.detail = data.get('detail', r.detail)
    db.session.commit()
    _upsert_chunk('requirement', r.id)
    return jsonify({'success': True})


@app.route('/api/requirements/<int:rid>', methods=['DELETE'])
def api_req_delete(rid):
    r = db.get_or_404(Requirement, rid)
    _delete_chunk('requirement', r.id)
    db.session.delete(r)
    db.session.commit()
    return jsonify({'success': True})


# ── Requirements + TOC Excel Export ──────────────────────────────────────────

@app.route('/api/pdf/<int:pdf_id>/export/excel')
def api_pdf_export_excel(pdf_id):
    pdf = db.get_or_404(PdfDocument, pdf_id)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = openpyxl.Workbook()

    hdr_fill = PatternFill('solid', fgColor='1E293B')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    left_al  = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ctr_al   = Alignment(horizontal='center', vertical='center')
    thin     = Side(style='thin', color='CBD5E1')
    bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_header(ws, headers, widths):
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = ctr_al; c.border = bdr
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = 'A2'

    # ── Sheet 1: 요구사항 목록 ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = '요구사항 목록'
    write_header(ws1,
        ['No', '고유번호', '명칭', '세부내용'],
        [5,     18,        40,     80])

    reqs = Requirement.query.filter_by(pdf_id=pdf_id).order_by(Requirement.id).all()
    for i, r in enumerate(reqs, 2):
        vals = [i - 1, r.req_id, r.req_name, r.detail or '']
        for col, val in enumerate(vals, 1):
            c = ws1.cell(row=i, column=col, value=val)
            c.border = bdr
            c.alignment = ctr_al if col == 1 else left_al
        ws1.row_dimensions[i].height = 40

    # ── Sheet 2: 목차 현황 ─────────────────────────────────────────
    ws2 = wb.create_sheet('목차 현황')
    write_header(ws2,
        ['No', '1단계', '2단계', '3단계', '작업상태', '쪽수', '연결 요구사항'],
        [5,     25,      25,      30,       11,          8,       30])

    TOC_COLORS = {
        '신규':   ('DBEAFE', '1D4ED8'),
        '수정필요': ('FEF3C7', '92400E'),
        '작업중':  ('F3E8FF', '6B21A8'),
        '완료':   ('DCFCE7', '166534'),
    }
    tocs = TOCItem.query.filter_by(pdf_id=pdf_id).order_by(TOCItem.order_index, TOCItem.id).all()
    for i, t in enumerate(tocs, 2):
        req_labels = ', '.join(r.req_id for r in t.requirements)
        vals = [i - 1, t.depth1 or '', t.depth2 or '', t.depth3 or '',
                t.work_status or '신규', t.page_number or '', req_labels]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=i, column=col, value=val)
            c.border = bdr
            c.alignment = ctr_al if col in (1, 5, 6) else left_al
        t_bg, t_fg = TOC_COLORS.get(t.work_status or '신규', ('F1F5F9', '475569'))
        ws2.cell(row=i, column=5).fill = PatternFill('solid', fgColor=t_bg)
        ws2.cell(row=i, column=5).font = Font(bold=True, color=t_fg, size=9)
        ws2.row_dimensions[i].height = 30

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    safe_name = pdf.original_name.rsplit('.', 1)[0] if '.' in pdf.original_name else pdf.original_name
    return send_file(
        output,
        as_attachment=True,
        download_name=f'{safe_name}_요구사항.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ── Dashboard API ─────────────────────────────────────────────────────────────

@app.route('/api/dashboard')
def api_dashboard():
    from datetime import date, timedelta
    today = date.today()
    soon  = today + timedelta(days=7)
    today_str = today.isoformat()
    soon_str  = soon.isoformat()

    # 마감 임박 할일 (오늘~7일 이내 + 이미 지난 것 포함)
    todos = TodoItem.query.filter(
        TodoItem.status != '완료',
        TodoItem.due_date != '',
        TodoItem.due_date != None,
        TodoItem.due_date <= soon_str,
    ).order_by(TodoItem.due_date).limit(10).all()

    # 최근 작업일지
    worklogs = WorkLog.query.order_by(WorkLog.created_at.desc()).limit(5).all()

    # 프로젝트별 TOC 완료율 (PDF별로 계산)
    projects = Project.query.order_by(Project.created_at.desc()).all()
    proj_stats = []
    for p in projects:
        # TOC가 있는 PDF 중 가장 최근 업로드(id 최대) 하나만 사용
        pdfs_with_toc = [pdf for pdf in p.pdfs if pdf.toc_items]
        pdf_toc = []
        if pdfs_with_toc:
            pdf = max(pdfs_with_toc, key=lambda x: x.id)
            items = pdf.toc_items
            leaf = [t for t in items if t.depth3] or items
            d1_done = {t.depth1 for t in items if not t.depth2 and not t.depth3 and t.work_status == '완료'}
            d2_done = {(t.depth1, t.depth2) for t in items if t.depth2 and not t.depth3 and t.work_status == '완료'}
            done = sum(
                1 for t in leaf
                if t.work_status == '완료'
                or t.depth1 in d1_done
                or (t.depth1, t.depth2) in d2_done
            )
            pdf_toc.append({
                'id':        pdf.id,
                'name':      pdf.original_name,
                'toc_total': len(leaf),
                'toc_done':  done,
            })
        req_total = sum(len(pdf.requirements) for pdf in p.pdfs)
        proj_stats.append({
            **p.to_dict(),
            'pdf_toc':  pdf_toc,
            'req_total': req_total,
        })

    return jsonify({
        'todo_upcoming': [t.to_dict() for t in todos],
        'worklog_recent': [w.to_dict() for w in worklogs],
        'projects': proj_stats,
        'today': today_str,
    })


# ── Requirement Proposal Image API ───────────────────────────────────────────

def _pick_option(value, allowed, default):
    return value if value in allowed else default


@app.route('/api/requirements/<int:rid>/proposal-images')
def api_req_proposal_images(rid):
    db.get_or_404(Requirement, rid)
    rows = RequirementProposalImage.query.filter_by(requirement_id=rid) \
        .order_by(RequirementProposalImage.created_at.desc(), RequirementProposalImage.id.desc()).all()
    return jsonify([r.to_dict() for r in rows])


@app.route('/api/requirements/<int:rid>/proposal-image', methods=['POST'])
def api_req_proposal_image_create(rid):
    req = db.get_or_404(Requirement, rid)
    data = request.get_json() or {}
    orientation = _pick_option(data.get('orientation'), {'landscape', 'portrait'}, 'landscape')
    template_type = _pick_option(data.get('template_type'), {'auto', 'function', 'architecture', 'checklist'}, 'auto')
    tone = _pick_option(data.get('tone'), {'public', 'technical', 'concise'}, 'public')

    from gemini_service import generate_requirement_proposal_content

    options = {
        'orientation': orientation,
        'template_type': template_type,
        'tone': tone,
    }
    content = generate_requirement_proposal_content(req.to_dict(), options, GEMINI_API_KEY)

    saved_filename = f'{uuid.uuid4().hex}.png'
    output_path = os.path.join(GENERATED_DIR, saved_filename)
    try:
        render_requirement_proposal_image(
            content=content,
            requirement=req.to_dict(),
            orientation=orientation,
            template_type=template_type,
            tone=tone,
            output_path=output_path,
        )
    except ProposalImageDependencyError as e:
        if os.path.exists(output_path):
            os.remove(output_path)
        return jsonify({'error': str(e)}), 500

    record = RequirementProposalImage(
        requirement_id=req.id,
        orientation=orientation,
        template_type=template_type,
        tone=tone,
        title=content.get('title') or f'{req.req_id} {req.req_name}',
        saved_filename=saved_filename,
        payload_json=json.dumps(content, ensure_ascii=False),
    )
    db.session.add(record)
    db.session.commit()
    return jsonify({'success': True, 'image': record.to_dict()})


@app.route('/api/proposal-images/<int:image_id>/file')
def api_proposal_image_file(image_id):
    image = db.get_or_404(RequirementProposalImage, image_id)
    path = os.path.join(GENERATED_DIR, image.saved_filename)
    if not os.path.exists(path):
        abort(404)
    return send_file(path, mimetype='image/png',
                     download_name=f'req_proposal_{image.id}.png')


# ── 요구사항 ↔ 목차 매핑 API ──────────────────────────────────────────────────

@app.route('/api/req/<int:req_id>/toc-mapping')
def api_req_toc_mapping(req_id):
    """이 요구사항이 연결된 목차 ID 목록 반환."""
    req = db.get_or_404(Requirement, req_id)
    mapped_ids = db.session.execute(
        db.select(toc_requirement.c.toc_id).where(toc_requirement.c.requirement_id == req_id)
    ).scalars().all()
    return jsonify(list(mapped_ids))


@app.route('/api/req/<int:req_id>/map-tocs', methods=['POST'])
def api_req_map_tocs(req_id):
    """요구사항을 여러 목차에 일괄 매핑 (전체 동기화)."""
    req = db.get_or_404(Requirement, req_id)
    new_toc_ids = set(request.json.get('toc_ids', []))
    toc_items = TOCItem.query.filter_by(pdf_id=req.pdf_id).all()
    for t in toc_items:
        has = req in t.requirements
        want = t.id in new_toc_ids
        if want and not has:
            t.requirements.append(req)
        elif not want and has:
            t.requirements.remove(req)
    db.session.commit()
    return jsonify({'ok': True, 'mapped': len(new_toc_ids)})


# ── TOC API ───────────────────────────────────────────────────────────────────

@app.route('/api/pdf/<int:pdf_id>/toc')
def api_toc(pdf_id):
    db.get_or_404(PdfDocument, pdf_id)
    items = TOCItem.query.filter_by(pdf_id=pdf_id).order_by(TOCItem.order_index, TOCItem.id).all()
    return jsonify([t.to_dict() for t in items])


@app.route('/api/pdf/<int:pdf_id>/toc', methods=['POST'])
def api_toc_create(pdf_id):
    db.get_or_404(PdfDocument, pdf_id)
    data = request.get_json()
    depth1 = data.get('depth1') or None
    depth2 = data.get('depth2') or None
    depth3 = data.get('depth3') or None

    # 삽입 위치 결정: 같은 부모 그룹의 마지막 항목 바로 뒤
    # depth3 → depth1+depth2 기준, depth2 → depth1 기준, depth1 → 맨 끝
    if depth1 and depth2 and depth3:
        # depth1+depth2가 같은 마지막 항목 뒤
        last_same = TOCItem.query.filter_by(pdf_id=pdf_id, depth1=depth1, depth2=depth2) \
            .order_by(TOCItem.order_index.desc()).first()
        if not last_same:
            # 해당 depth2 그룹이 없으면 depth1 마지막 뒤
            last_same = TOCItem.query.filter_by(pdf_id=pdf_id, depth1=depth1) \
                .order_by(TOCItem.order_index.desc()).first()
    elif depth1 and depth2:
        last_same = TOCItem.query.filter_by(pdf_id=pdf_id, depth1=depth1) \
            .order_by(TOCItem.order_index.desc()).first()
    else:
        last_same = None

    if last_same:
        insert_pos = last_same.order_index + 1
        for item in TOCItem.query.filter(
            TOCItem.pdf_id == pdf_id,
            TOCItem.order_index >= insert_pos
        ).all():
            item.order_index += 1
    else:
        insert_pos = TOCItem.query.filter_by(pdf_id=pdf_id).count()

    t = TOCItem(
        pdf_id=pdf_id,
        depth1=depth1,
        depth2=depth2,
        depth3=depth3,
        remarks=data.get('remarks') or None,
        work_status=data.get('work_status', '신규'),
        page_number=data.get('page_number') or None,
        order_index=insert_pos,
    )
    for rid in (data.get('requirement_ids') or []):
        req = db.session.get(Requirement, int(rid))
        if req and req.pdf_id == pdf_id:
            t.requirements.append(req)
    db.session.add(t)
    db.session.commit()
    return jsonify({'success': True, 'id': t.id})


def _propagate_completion(t):
    """depth3 → depth2 → depth1 순으로 완료 상태를 전파한다."""
    # depth3 항목이 완료됐을 때 → depth2 부모 확인
    if t.depth3 and t.depth2 and t.depth1:
        siblings = TOCItem.query.filter_by(
            pdf_id=t.pdf_id, depth1=t.depth1, depth2=t.depth2
        ).filter(TOCItem.depth3.isnot(None)).all()
        if siblings and all(s.work_status == '완료' for s in siblings):
            parent = TOCItem.query.filter_by(
                pdf_id=t.pdf_id, depth1=t.depth1, depth2=t.depth2, depth3=None
            ).first()
            if parent and parent.work_status != '완료':
                parent.work_status = '완료'
                _propagate_completion(parent)  # depth2 → depth1 전파

    # depth2 항목이 완료됐을 때 → depth1 부모 확인
    elif t.depth2 and t.depth1 and not t.depth3:
        siblings = TOCItem.query.filter_by(
            pdf_id=t.pdf_id, depth1=t.depth1
        ).filter(TOCItem.depth2.isnot(None)).all()
        if siblings and all(s.work_status == '완료' for s in siblings):
            parent = TOCItem.query.filter_by(
                pdf_id=t.pdf_id, depth1=t.depth1, depth2=None, depth3=None
            ).first()
            if parent and parent.work_status != '완료':
                parent.work_status = '완료'


@app.route('/api/toc/<int:tid>', methods=['PUT'])
def api_toc_update(tid):
    t = db.get_or_404(TOCItem, tid)
    data = request.get_json()
    # 부분 업데이트: 요청에 포함된 필드만 갱신 (없는 필드는 기존값 유지)
    if 'depth1' in data: t.depth1 = data['depth1'] or None
    if 'depth2' in data: t.depth2 = data['depth2'] or None
    if 'depth3' in data: t.depth3 = data['depth3'] or None
    if 'remarks' in data: t.remarks = data['remarks'] or None
    if 'work_status' in data: t.work_status = data['work_status']
    if 'page_number' in data: t.page_number = data['page_number'] or None
    if 'requirement_ids' in data:
        t.requirements = []
        for rid in (data['requirement_ids'] or []):
            req = db.session.get(Requirement, int(rid))
            if req and req.pdf_id == t.pdf_id:
                t.requirements.append(req)
    db.session.commit()
    # 완료 상태 상위 전파
    if t.work_status == '완료':
        _propagate_completion(t)
        db.session.commit()
    if t.depth3:   # depth3 항목만 청크 대상
        _upsert_chunk('toc', t.id)
    return jsonify({'success': True})


@app.route('/api/toc/<int:tid>', methods=['DELETE'])
def api_toc_delete(tid):
    t = db.get_or_404(TOCItem, tid)
    _delete_chunk('toc', t.id)
    db.session.delete(t)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/pdf/<int:pdf_id>/toc/reorder', methods=['PUT'])
def api_toc_reorder(pdf_id):
    db.get_or_404(PdfDocument, pdf_id)
    order = request.get_json()
    id_map = {item['id']: item['order_index'] for item in order}
    items = TOCItem.query.filter(
        TOCItem.pdf_id == pdf_id,
        TOCItem.id.in_(id_map.keys())
    ).all()
    for t in items:
        t.order_index = id_map[t.id]
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/pdf/<int:pdf_id>/toc/normalize', methods=['POST'])
def api_toc_normalize(pdf_id):
    """같은 depth1 항목이 연속되도록 순서를 재정리하고, depth3 항목의 depth2를 위치 기반으로 정정."""
    db.get_or_404(PdfDocument, pdf_id)
    items = TOCItem.query.filter_by(pdf_id=pdf_id).order_by(TOCItem.order_index, TOCItem.id).all()

    # depth1별 그룹화 (헤더 우선, 순서는 첫 등장 순)
    from collections import OrderedDict
    groups = OrderedDict()
    for t in items:
        key = t.depth1 or ''
        if key not in groups:
            groups[key] = []
        groups[key].append(t)

    # 재정렬: 각 그룹 내 헤더(depth2/3 없는 것) → 나머지 순
    ordered = []
    idx = 0
    for key, group in groups.items():
        headers = [t for t in group if not t.depth2 and not t.depth3]
        children = [t for t in group if t.depth2 or t.depth3]
        for t in headers + children:
            t.order_index = idx
            idx += 1
            ordered.append(t)

    # depth3 항목의 depth2를 위치 기반으로 정정
    # 같은 depth1 내에서 직전 depth2 항목(depth2 있고 depth3 없는)의 depth2 값을 사용
    current_d1 = None
    current_d2 = None
    for t in ordered:
        if t.depth1 != current_d1:
            current_d1 = t.depth1
            current_d2 = None
        if t.depth2 and not t.depth3:
            current_d2 = t.depth2
        elif t.depth3 and current_d2 and t.depth2 != current_d2:
            t.depth2 = current_d2

    db.session.commit()
    return jsonify({'success': True})


# ── Matrix API ────────────────────────────────────────────────────────────────

@app.route('/api/pdf/<int:pdf_id>/matrix')
def api_matrix(pdf_id):
    db.get_or_404(PdfDocument, pdf_id)

    reqs = Requirement.query.filter_by(pdf_id=pdf_id).order_by(Requirement.id).all()
    tocs = TOCItem.query.filter_by(pdf_id=pdf_id).order_by(TOCItem.order_index, TOCItem.id).all()

    linked_req_ids = set()
    req_toc_map = {r.id: [] for r in reqs}
    toc_rows = []
    for t in tocs:
        req_ids = [r.id for r in t.requirements]
        linked_req_ids.update(req_ids)
        toc_path = ' > '.join(filter(None, [t.depth1, t.depth2, t.depth3]))
        for r in t.requirements:
            if r.id in req_toc_map:
                req_toc_map[r.id].append({
                    'id': t.id,
                    'depth1': t.depth1 or '',
                    'depth2': t.depth2 or '',
                    'depth3': t.depth3 or '',
                    'path': toc_path,
                    'page_number': t.page_number or '',
                    'work_status': t.work_status or '신규',
                    'order_index': t.order_index,
                })
        toc_rows.append({
            'id': t.id,
            'depth1': t.depth1 or '',
            'depth2': t.depth2 or '',
            'depth3': t.depth3 or '',
            'work_status': t.work_status,
            'page_number': t.page_number or '',
            'requirement_ids': req_ids,
            'requirement_labels': [r.req_id for r in t.requirements],
        })

    requirement_rows = []
    for r in reqs:
        mapped_tocs = sorted(req_toc_map.get(r.id, []), key=lambda x: (x['order_index'], x['id']))
        requirement_rows.append({
            **r.to_dict(),
            'toc_items': mapped_tocs,
            'toc_count': len(mapped_tocs),
        })

    unlinked = [r.to_dict() for r in reqs if r.id not in linked_req_ids]

    return jsonify({
        'toc_rows': toc_rows,
        'requirement_rows': requirement_rows,
        'unlinked_requirements': unlinked,
        'total_req': len(reqs),
        'linked_count': len(linked_req_ids),
        'unlinked_count': len(unlinked),
    })


# ── Search API ────────────────────────────────────────────────────────────────

@app.route('/api/search')
def api_search():
    q = (request.args.get('q') or '').strip()
    project_id = request.args.get('project_id', type=int)

    if not q or len(q) < 2:
        return jsonify({'results': [], 'total': 0})

    query = Requirement.query.join(PdfDocument).join(Project)
    if project_id:
        query = query.filter(PdfDocument.project_id == project_id)

    like = f'%{q}%'
    query = query.filter(
        or_(
            Requirement.req_id.ilike(like),
            Requirement.req_name.ilike(like),
            Requirement.detail.ilike(like),
        )
    ).order_by(Project.id, PdfDocument.id, Requirement.id).limit(200)

    rows = query.all()

    # Group by project > pdf
    groups: dict = {}
    for r in rows:
        pdf = r.pdf
        proj = pdf.project
        pk = proj.id
        fk = pdf.id
        if pk not in groups:
            groups[pk] = {'project_id': pk, 'project_name': proj.name, 'pdfs': {}}
        if fk not in groups[pk]['pdfs']:
            groups[pk]['pdfs'][fk] = {
                'pdf_id': fk,
                'project_id': pk,
                'original_name': pdf.original_name,
                'requirements': [],
            }
        groups[pk]['pdfs'][fk]['requirements'].append({
            **r.to_dict(),
            'project_name': proj.name,
        })

    results = []
    for pg in groups.values():
        for pf in pg['pdfs'].values():
            results.append({
                'project_id': pg['project_id'],
                'project_name': pg['project_name'],
                'pdf_id': pf['pdf_id'],
                'original_name': pf['original_name'],
                'requirements': pf['requirements'],
            })

    return jsonify({'results': results, 'total': len(rows)})


@app.route('/api/projects/list')
def api_projects_list():
    projects = Project.query.order_by(Project.name).all()
    return jsonify([{'id': p.id, 'name': p.name} for p in projects])


# ── 요구사항 타입 규칙 API ────────────────────────────────────────────────────────

@app.route('/settings/req-types')
def req_types_page():
    return render_template('req_types.html')


@app.route('/api/req-types')
def api_req_types_list():
    rules = ReqTypeRule.query.order_by(ReqTypeRule.order_index, ReqTypeRule.id).all()
    return jsonify([r.to_dict() for r in rules])


@app.route('/api/req-types', methods=['POST'])
def api_req_types_create():
    data = request.get_json()
    prefix = (data.get('prefix') or '').strip().upper()
    label = (data.get('label') or '').strip()
    if not prefix or not label:
        return jsonify({'error': 'prefix와 label은 필수입니다.'}), 400
    count = ReqTypeRule.query.count()
    rule = ReqTypeRule(
        prefix=prefix, label=label,
        bg_color=data.get('bg_color', '#dbeafe'),
        text_color=data.get('text_color', '#1d4ed8'),
        order_index=count,
    )
    db.session.add(rule)
    db.session.commit()
    return jsonify({'success': True, 'rule': rule.to_dict()})


@app.route('/api/req-types/<int:rid>', methods=['PUT'])
def api_req_types_update(rid):
    rule = db.get_or_404(ReqTypeRule, rid)
    data = request.get_json()
    if 'prefix' in data:
        rule.prefix = (data['prefix'] or '').strip().upper()
    if 'label' in data:
        rule.label = (data['label'] or '').strip()
    if 'bg_color' in data:
        rule.bg_color = data['bg_color']
    if 'text_color' in data:
        rule.text_color = data['text_color']
    db.session.commit()
    return jsonify({'success': True, 'rule': rule.to_dict()})


@app.route('/api/req-types/<int:rid>', methods=['DELETE'])
def api_req_types_delete(rid):
    rule = db.get_or_404(ReqTypeRule, rid)
    db.session.delete(rule)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/req-types/reorder', methods=['PUT'])
def api_req_types_reorder():
    order = request.get_json()
    id_map = {item['id']: item['order_index'] for item in order}
    rules = ReqTypeRule.query.filter(ReqTypeRule.id.in_(id_map.keys())).all()
    for r in rules:
        r.order_index = id_map[r.id]
    db.session.commit()
    return jsonify({'success': True})


# ── Proposal Match Analysis ───────────────────────────────────────────────────

def _run_analysis_db(task_id: str, proposal_path: str, proposal_name: str,
                     requirements: list, toc_items: list, rfp_label: str):
    """DB 요구사항 기반 분석 — 제안서 PDF만 Gemini에 업로드."""
    from gemini_service import analyze_proposal_with_db_requirements, generate_action_summary
    q = _task_queues[task_id]

    def send(step: int, msg: str):
        logger.info("[analyze_db:%s] step %d: %s", task_id[:8], step, msg)
        q.put({'type': 'progress', 'step': step, 'total': 5, 'msg': msg})

    try:
        # Steps 1-3: DB 요구사항 기반 분석 (RFP 업로드 없음)
        results = analyze_proposal_with_db_requirements(
            proposal_path, requirements, toc_items, GEMINI_API_KEY,
            progress_cb=send,
        )

        # Step 4: 종합 요약
        send(4, '종합 액션 플랜 생성 중...')
        summary = generate_action_summary(results, GEMINI_API_KEY)

        # Step 5: DB 저장
        send(5, 'DB에 저장 중...')
        mappings  = results.get('proposal_mappings', [])
        uncovered = results.get('uncovered_rfp', [])
        full    = sum(1 for r in mappings if r.get('match_level') == '완전')
        partial = sum(1 for r in mappings if r.get('match_level') == '부분')
        missing = len(uncovered)
        total   = full + partial + missing

        with app.app_context():
            record = ProposalAnalysis(
                rfp_name=rfp_label,
                proposal_name=proposal_name,
                results_json=json.dumps(results, ensure_ascii=False),
                summary_json=json.dumps(summary, ensure_ascii=False),
                total_count=total,
                full_count=full,
                partial_count=partial,
                missing_count=missing,
            )
            db.session.add(record)
            db.session.commit()
            record_id = record.id

        _analyze_results[task_id] = {'record_id': record_id, 'results': results, 'summary': summary}
        q.put({'type': 'done', 'record_id': record_id,
               'msg': f'매핑 {len(mappings)}건 / 미충족 {len(uncovered)}건'})
    except Exception as e:
        logger.error("[analyze_db:%s] 실패: %s", task_id[:8], e)
        q.put({'type': 'error', 'msg': str(e)})
    finally:
        try:
            os.remove(proposal_path)
        except Exception:
            pass


def _run_analysis(task_id: str, rfp_path: str, proposal_path: str,
                  rfp_name: str, proposal_name: str):
    from gemini_service import analyze_proposal_match, generate_action_summary
    q = _task_queues[task_id]

    def send(step: int, msg: str):
        logger.info("[analyze:%s] step %d: %s", task_id[:8], step, msg)
        q.put({'type': 'progress', 'step': step, 'total': 6, 'msg': msg})

    try:
        # Steps 1-4: 매칭 분석
        results = analyze_proposal_match(
            rfp_path, proposal_path, GEMINI_API_KEY,
            progress_cb=send,
        )

        # Step 5: 종합 요약 생성
        send(5, '종합 액션 플랜 생성 중...')
        summary = generate_action_summary(results, GEMINI_API_KEY)

        # Step 6: DB 저장
        send(6, 'DB에 저장 중...')
        mappings = results.get('proposal_mappings', [])
        uncovered = results.get('uncovered_rfp', [])
        full    = sum(1 for r in mappings if r.get('match_level') == '완전')
        partial = sum(1 for r in mappings if r.get('match_level') == '부분')
        missing = len(uncovered)
        total   = full + partial + missing

        with app.app_context():
            record = ProposalAnalysis(
                rfp_name=rfp_name,
                proposal_name=proposal_name,
                results_json=json.dumps(results, ensure_ascii=False),
                summary_json=json.dumps(summary, ensure_ascii=False),
                total_count=total,
                full_count=full,
                partial_count=partial,
                missing_count=missing,
            )
            db.session.add(record)
            db.session.commit()
            record_id = record.id

        _analyze_results[task_id] = {'record_id': record_id, 'results': results, 'summary': summary}
        msg = (f'제안서 매핑 {len(results.get("proposal_mappings",[]))}건 / '
               f'미충족 RFP {len(results.get("uncovered_rfp",[]))}건')
        q.put({'type': 'done', 'record_id': record_id, 'msg': msg})

    except Exception as e:
        logger.error("[analyze:%s] 실패: %s", task_id[:8], e)
        q.put({'type': 'error', 'msg': str(e)})
    finally:
        for p in [rfp_path, proposal_path]:
            try:
                os.remove(p)
            except Exception:
                pass


@app.route('/analyze')
def analyze_page():
    return render_template('analyze.html')


@app.route('/api/analyze/db-sources')
def api_analyze_db_sources():
    """DB에 저장된 프로젝트/PDF/요구사항 목록을 반환."""
    projects = Project.query.order_by(Project.created_at.desc()).all()
    result = []
    for proj in projects:
        pdfs = []
        for pdf in sorted(proj.pdfs, key=lambda p: p.created_at or 0, reverse=True):
            reqs = sorted(pdf.requirements, key=lambda r: r.id)
            if not reqs:
                continue
            pdfs.append({
                'id':   pdf.id,
                'name': pdf.original_name,
                'requirements': [
                    {'id': r.id, 'req_id': r.req_id, 'req_name': r.req_name}
                    for r in reqs
                ],
            })
        if pdfs:
            result.append({'id': proj.id, 'name': proj.name, 'pdfs': pdfs})
    return jsonify(result)


@app.route('/api/analyze/run-db', methods=['POST'])
def api_analyze_run_db():
    """DB 요구사항 기반 분석 — 제안서 PDF만 업로드."""
    proposal_file = request.files.get('proposal')
    req_ids_raw   = request.form.get('req_ids', '')
    rfp_label     = request.form.get('rfp_label', 'DB 요구사항')

    if not proposal_file:
        return jsonify({'error': '제안서 파일을 업로드하세요.'}), 400
    if not proposal_file.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'PDF 파일만 허용됩니다.'}), 400
    if not req_ids_raw:
        return jsonify({'error': 'DB에서 하나 이상의 요구사항을 선택하세요.'}), 400

    try:
        req_ids = [int(x) for x in req_ids_raw.split(',') if x.strip()]
    except ValueError:
        return jsonify({'error': '잘못된 요구사항 ID 형식입니다.'}), 400

    # 선택된 요구사항 로드
    requirements, toc_items = [], []
    seen_pdf_ids = set()
    for req_id in req_ids:
        r = db.session.get(Requirement, req_id)
        if not r:
            continue
        requirements.append({'req_id': r.req_id, 'req_name': r.req_name, 'detail': r.detail or ''})
        if r.pdf_id not in seen_pdf_ids:
            seen_pdf_ids.add(r.pdf_id)
            for t in r.pdf.toc_items:
                toc_items.append({'depth1': t.depth1 or '', 'depth2': t.depth2 or '', 'depth3': t.depth3 or ''})

    if not requirements:
        return jsonify({'error': '선택한 요구사항을 찾을 수 없습니다.'}), 400

    proposal_name = proposal_file.filename
    proposal_path = os.path.join(app.config['UPLOAD_FOLDER'], f"prop_{uuid.uuid4().hex}.pdf")
    proposal_file.save(proposal_path)

    task_id = uuid.uuid4().hex
    _task_queues[task_id] = queue.Queue()
    threading.Thread(
        target=_run_analysis_db,
        args=(task_id, proposal_path, proposal_name, requirements, toc_items, rfp_label),
        daemon=True,
    ).start()
    return jsonify({'task_id': task_id, 'req_count': len(requirements)})


@app.route('/api/analyze/run', methods=['POST'])
def api_analyze_run():
    rfp_file      = request.files.get('rfp')
    proposal_file = request.files.get('proposal')
    if not rfp_file or not proposal_file:
        return jsonify({'error': '제안요청서(RFP)와 제안서 파일을 모두 업로드하세요.'}), 400
    if not rfp_file.filename.lower().endswith('.pdf') or not proposal_file.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'PDF 파일만 허용됩니다.'}), 400

    rfp_name      = rfp_file.filename
    proposal_name = proposal_file.filename
    rfp_path      = os.path.join(app.config['UPLOAD_FOLDER'], f"rfp_{uuid.uuid4().hex}.pdf")
    proposal_path = os.path.join(app.config['UPLOAD_FOLDER'], f"prop_{uuid.uuid4().hex}.pdf")
    rfp_file.save(rfp_path)
    proposal_file.save(proposal_path)

    task_id = uuid.uuid4().hex
    _task_queues[task_id] = queue.Queue()
    threading.Thread(
        target=_run_analysis,
        args=(task_id, rfp_path, proposal_path, rfp_name, proposal_name),
        daemon=True,
    ).start()
    return jsonify({'task_id': task_id})


@app.route('/api/analyze/progress/<task_id>')
def api_analyze_progress(task_id):
    def generate():
        q = _task_queues.get(task_id)
        if not q:
            yield f"data: {json.dumps({'type': 'error', 'msg': '작업을 찾을 수 없습니다.'})}\n\n"
            return
        try:
            while True:
                try:
                    msg = q.get(timeout=180)
                    yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"
                    if msg.get('type') in ('done', 'error'):
                        break
                except queue.Empty:
                    yield 'data: {"type":"keepalive"}\n\n'
        finally:
            _task_queues.pop(task_id, None)

    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'},
    )


@app.route('/api/analyze/result/<task_id>')
def api_analyze_result(task_id):
    data = _analyze_results.get(task_id)
    if data is None:
        return jsonify({'error': 'result not ready'}), 404
    return jsonify(data)


@app.route('/api/analyze/history')
def api_analyze_history():
    records = ProposalAnalysis.query.order_by(ProposalAnalysis.created_at.desc()).limit(50).all()
    return jsonify([r.to_dict() for r in records])


@app.route('/api/analyze/history/<int:record_id>')
def api_analyze_history_detail(record_id):
    record = db.get_or_404(ProposalAnalysis, record_id)
    return jsonify(record.to_full_dict())


@app.route('/api/analyze/history/<int:record_id>', methods=['DELETE'])
def api_analyze_history_delete(record_id):
    record = db.get_or_404(ProposalAnalysis, record_id)
    db.session.delete(record)
    db.session.commit()
    return jsonify({'success': True})


def _build_excel(results: list, summary: dict, rfp_name: str, proposal_name: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = openpyxl.Workbook()

    # ── Sheet 1: 매칭 결과 ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = '요구사항 매칭 분석'

    hdr_fill  = PatternFill('solid', fgColor='1E293B')
    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    center_al = Alignment(horizontal='center', vertical='top', wrap_text=True)
    left_al   = Alignment(horizontal='left',   vertical='top', wrap_text=True)
    thin      = Side(style='thin', color='CBD5E1')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_map  = {'완전': PatternFill('solid', fgColor='DCFCE7'),
                 '부분': PatternFill('solid', fgColor='FEF3C7'),
                 '미반영': PatternFill('solid', fgColor='FEE2E2')}
    font_map  = {'완전': Font(bold=True, color='166534', size=9),
                 '부분': Font(bold=True, color='92400E', size=9),
                 '미반영': Font(bold=True, color='991B1B', size=9)}

    mappings  = results.get('proposal_mappings', []) if isinstance(results, dict) else results
    uncovered = results.get('uncovered_rfp', [])     if isinstance(results, dict) else []

    # ── 제안서 → RFP 매핑 시트 ───────────────────────────────────────────────
    headers    = ['No', '제안서 섹션', '제안서 내용', 'RFP 섹션', 'RFP 요구사항', '매칭 수준', '매칭 근거']
    col_widths = [5, 22, 42, 22, 42, 10, 55]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = center_al; cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    for row_idx, item in enumerate(mappings, 2):
        level  = item.get('match_level', '')
        values = [item.get('no', row_idx - 1),
                  item.get('proposal_section', ''), item.get('proposal_content', ''),
                  item.get('rfp_section', ''),      item.get('rfp_requirement', ''),
                  level, item.get('match_reason', '')]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = center_al if col in (1, 6) else left_al
            cell.border = border
            if col == 6 and level in fill_map:
                cell.fill = fill_map[level]; cell.font = font_map[level]
    ws.freeze_panes = 'A2'

    # ── 미충족 RFP 요구사항 시트 ─────────────────────────────────────────────
    if uncovered:
        ws_gap = wb.create_sheet('미충족 RFP 요구사항')
        gap_headers    = ['No', 'RFP 섹션', 'RFP 요구사항', '상태', '추가 위치 (제안서)', '추가 내용 제안']
        gap_col_widths = [5, 22, 42, 10, 28, 55]
        for col, (h, w) in enumerate(zip(gap_headers, gap_col_widths), 1):
            cell = ws_gap.cell(row=1, column=col, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = center_al; cell.border = border
            ws_gap.column_dimensions[get_column_letter(col)].width = w
        ws_gap.row_dimensions[1].height = 22
        for row_idx, item in enumerate(uncovered, 2):
            status = item.get('status', '미반영')
            values = [item.get('no', row_idx - 1),
                      item.get('rfp_section', ''), item.get('rfp_requirement', ''),
                      status,
                      item.get('add_to_section', ''), item.get('suggestion', '')]
            for col, val in enumerate(values, 1):
                cell = ws_gap.cell(row=row_idx, column=col, value=val)
                cell.alignment = center_al if col in (1, 4) else left_al
                cell.border = border
                if col == 4:
                    cell.fill = fill_map.get(status, fill_map['미반영'])
                    cell.font = font_map.get(status, font_map['미반영'])
            ws_gap.row_dimensions[row_idx].height = 50
        ws_gap.freeze_panes = 'A2'

    # ── Sheet 2: 액션 플랜 ────────────────────────────────────────────────────
    if summary:
        ws2 = wb.create_sheet('액션 플랜')
        r = 1

        def write_hdr(text, color='1E293B'):
            c = ws2.cell(row=r, column=1, value=text)
            c.font = Font(bold=True, color='FFFFFF', size=11)
            c.fill = PatternFill('solid', fgColor=color)
            c.alignment = Alignment(vertical='center')
            ws2.merge_cells(f'A{r}:E{r}')
            ws2.row_dimensions[r].height = 20

        def write_row(*vals, bold=False, bg=None):
            for ci, v in enumerate(vals, 1):
                c = ws2.cell(row=r, column=ci, value=v)
                c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                c.border = border
                if bold:
                    c.font = Font(bold=True, size=9)
                if bg:
                    c.fill = PatternFill('solid', fgColor=bg)

        # 기본 정보
        ws2.cell(row=r, column=1, value=f'RFP: {rfp_name}').font = Font(bold=True)
        r += 1
        ws2.cell(row=r, column=1, value=f'제안서: {proposal_name}').font = Font(bold=True)
        r += 2

        score = summary.get('overall_score', '')
        ws2.cell(row=r, column=1, value=f'전체 충족도: {score}점').font = Font(bold=True, size=12, color='1D4ED8')
        r += 1
        ws2.cell(row=r, column=1, value=summary.get('overall_assessment', ''))
        ws2.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
        ws2.row_dimensions[r].height = 60
        r += 2

        # 추가 필요
        write_hdr('➕ 추가 필요 항목', '991B1B'); r += 1
        write_row('우선순위', 'RFP 섹션', '요구사항', '작성 제안', bold=True, bg='FEE2E2'); r += 1
        for item in summary.get('add_items', []):
            write_row(item.get('priority', ''), item.get('rfp_section', ''),
                      item.get('requirement', ''), item.get('suggestion', ''))
            ws2.row_dimensions[r].height = 50; r += 1
        r += 1

        # 보완 필요
        write_hdr('✏️ 보완 필요 항목', '92400E'); r += 1
        write_row('우선순위', 'RFP 섹션', '제안서 섹션', '현재 부족한 점', '보완 방법', bold=True, bg='FEF3C7'); r += 1
        for item in summary.get('improve_items', []):
            write_row(item.get('priority', ''), item.get('rfp_section', ''),
                      item.get('proposal_section', ''), item.get('issue', ''), item.get('suggestion', ''))
            ws2.row_dimensions[r].height = 50; r += 1
        r += 1

        # 제거 권장
        write_hdr('🗑️ 제거/축소 권장', '374151'); r += 1
        write_row('제안서 섹션', '이유', bold=True, bg='F1F5F9'); r += 1
        for item in summary.get('remove_items', []):
            write_row(item.get('proposal_section', ''), item.get('reason', ''))
            ws2.row_dimensions[r].height = 40; r += 1
        r += 1

        # 우선순위 액션
        write_hdr('⚡ 우선순위 액션', '1D4ED8'); r += 1
        for action in summary.get('priority_actions', []):
            ws2.cell(row=r, column=1, value=action)
            ws2.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
            ws2.merge_cells(f'A{r}:E{r}')
            ws2.row_dimensions[r].height = 35; r += 1

        ws2.column_dimensions['A'].width = 10
        ws2.column_dimensions['B'].width = 25
        ws2.column_dimensions['C'].width = 25
        ws2.column_dimensions['D'].width = 40
        ws2.column_dimensions['E'].width = 40

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route('/api/analyze/download/<task_id>')
def api_analyze_download(task_id):
    data = _analyze_results.get(task_id)
    if not data:
        return jsonify({'error': 'result not found'}), 404
    results  = data.get('results', [])
    summary  = data.get('summary', {})
    record   = db.session.get(ProposalAnalysis, data.get('record_id'))
    rfp_name      = record.rfp_name      if record else 'RFP'
    proposal_name = record.proposal_name if record else '제안서'

    output = _build_excel(results, summary, rfp_name, proposal_name)
    return send_file(
        output,
        as_attachment=True,
        download_name='요구사항_매칭_분석.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/api/analyze/download/record/<int:record_id>')
def api_analyze_download_record(record_id):
    record = db.get_or_404(ProposalAnalysis, record_id)
    results = json.loads(record.results_json) if record.results_json else []
    summary = json.loads(record.summary_json) if record.summary_json else {}
    output  = _build_excel(results, summary, record.rfp_name, record.proposal_name)
    return send_file(
        output,
        as_attachment=True,
        download_name='요구사항_매칭_분석.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ── Revision Verification ─────────────────────────────────────────────────────

def _run_verify(task_id: str, revised_path: str, gap_items: list):
    from gemini_service import verify_revision
    q = _task_queues[task_id]

    def send(step: int, msg: str):
        logger.info("[verify:%s] step %d: %s", task_id[:8], step, msg)
        q.put({'type': 'progress', 'step': step, 'total': 3, 'msg': msg})

    try:
        results = verify_revision(revised_path, gap_items, GEMINI_API_KEY, progress_cb=send)
        _verify_results[task_id] = results
        resolved = sum(1 for r in results if r.get('resolved'))
        q.put({'type': 'done', 'msg': f'{resolved}/{len(results)}건 해결 확인'})
    except Exception as e:
        logger.error("[verify:%s] 실패: %s", task_id[:8], e)
        q.put({'type': 'error', 'msg': str(e)})
    finally:
        try:
            os.remove(revised_path)
        except Exception:
            pass


@app.route('/api/analyze/verify/<int:record_id>', methods=['POST'])
def api_analyze_verify(record_id):
    record = db.get_or_404(ProposalAnalysis, record_id)
    revised = request.files.get('revised_proposal')
    if not revised:
        return jsonify({'error': '수정된 제안서 파일을 업로드하세요.'}), 400
    if not revised.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'PDF 파일만 허용됩니다.'}), 400

    results_data = json.loads(record.results_json) if record.results_json else {}
    if isinstance(results_data, list):
        gap_items = []
    else:
        gap_items = results_data.get('uncovered_rfp', [])

    if not gap_items:
        return jsonify({'error': '이전 분석에 미충족 항목이 없습니다.'}), 400

    revised_path = os.path.join(app.config['UPLOAD_FOLDER'], f"rev_{uuid.uuid4().hex}.pdf")
    revised.save(revised_path)

    task_id = uuid.uuid4().hex
    _task_queues[task_id] = queue.Queue()
    threading.Thread(
        target=_run_verify,
        args=(task_id, revised_path, gap_items),
        daemon=True,
    ).start()
    return jsonify({'task_id': task_id, 'gap_count': len(gap_items)})


@app.route('/api/analyze/verify/progress/<task_id>')
def api_verify_progress(task_id):
    def generate():
        q = _task_queues.get(task_id)
        if not q:
            yield f"data: {json.dumps({'type':'error','msg':'작업을 찾을 수 없습니다.'})}\n\n"
            return
        try:
            while True:
                try:
                    msg = q.get(timeout=180)
                    yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"
                    if msg.get('type') in ('done', 'error'):
                        break
                except queue.Empty:
                    yield 'data: {"type":"keepalive"}\n\n'
        finally:
            _task_queues.pop(task_id, None)

    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'},
    )


@app.route('/api/analyze/verify/result/<task_id>')
def api_verify_result(task_id):
    results = _verify_results.get(task_id)
    if results is None:
        return jsonify({'error': 'result not ready'}), 404
    return jsonify(results)


# ══════════════════════════════════════════════════════════════════════════════
# 할일 목록 (Todo)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/calendar')
def calendar_page():
    return render_template('calendar.html')


@app.route('/todos')
def todos_page():
    return render_template('todos.html')


@app.route('/api/todos', methods=['GET'])
def api_todos_list():
    items = TodoItem.query.order_by(TodoItem.status, TodoItem.order_index).all()
    return jsonify([t.to_dict() for t in items])


@app.route('/api/todos', methods=['POST'])
def api_todos_create():
    data = request.get_json()
    from datetime import datetime as dt
    t = TodoItem(
        title=data['title'],
        description=data.get('description', ''),
        priority=data.get('priority', '보통'),
        status=data.get('status', '할일'),
        start_date=data.get('start_date', ''),
        due_date=data.get('due_date', ''),
        order_index=TodoItem.query.filter_by(status=data.get('status', '할일')).count(),
        created_at=dt.utcnow(),
        updated_at=dt.utcnow(),
    )
    db.session.add(t)
    db.session.commit()
    _upsert_chunk('todo', t.id)
    return jsonify(t.to_dict()), 201


@app.route('/api/todos/<int:todo_id>', methods=['PUT'])
def api_todos_update(todo_id):
    t = db.session.get(TodoItem, todo_id) or abort(404)
    data = request.get_json()
    from datetime import datetime as dt
    for field in ('title', 'description', 'priority', 'status', 'start_date', 'due_date', 'order_index'):
        if field in data:
            setattr(t, field, data[field])
    t.updated_at = dt.utcnow()
    db.session.commit()
    # 텍스트 변경이 있을 때만 재임베딩
    if any(f in data for f in ('title', 'description', 'priority', 'status', 'start_date', 'due_date')):
        _upsert_chunk('todo', t.id)
    return jsonify(t.to_dict())


@app.route('/api/todos/<int:todo_id>', methods=['DELETE'])
def api_todos_delete(todo_id):
    t = db.session.get(TodoItem, todo_id) or abort(404)
    _delete_chunk('todo', t.id)
    db.session.delete(t)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/todos/reorder', methods=['POST'])
def api_todos_reorder():
    """[{id, status, order_index}, ...] 배열로 일괄 업데이트"""
    from datetime import datetime as dt
    items = request.get_json()
    for item in items:
        t = db.session.get(TodoItem, item['id'])
        if t:
            t.status = item['status']
            t.order_index = item['order_index']
            t.updated_at = dt.utcnow()
    db.session.commit()
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════════════════════════════
# 작업일지 (WorkLog)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/worklog')
def worklog_page():
    return render_template('worklog.html')


@app.route('/api/worklogs', methods=['GET'])
def api_worklogs_list():
    logs = WorkLog.query.order_by(WorkLog.created_at.desc()).all()
    return jsonify([l.to_dict() for l in logs])


@app.route('/api/worklogs/<int:log_id>', methods=['GET'])
def api_worklog_get(log_id):
    log = db.session.get(WorkLog, log_id) or abort(404)
    return jsonify(log.to_dict(include_content=True))


@app.route('/api/worklogs', methods=['POST'])
def api_worklogs_create():
    data = request.get_json()
    from datetime import datetime as dt
    now = dt.utcnow()
    log = WorkLog(
        title=data['title'],
        content=data.get('content', ''),
        tags=json.dumps(data.get('tags', []), ensure_ascii=False),
        created_at=now,
        updated_at=now,
    )
    db.session.add(log)
    db.session.commit()
    _upsert_chunk('worklog', log.id)
    return jsonify(log.to_dict(include_content=True)), 201


@app.route('/api/worklogs/<int:log_id>', methods=['PUT'])
def api_worklog_update(log_id):
    log = db.session.get(WorkLog, log_id) or abort(404)
    data = request.get_json()
    from datetime import datetime as dt
    if 'title' in data:
        log.title = data['title']
    if 'content' in data:
        log.content = data['content']
    if 'tags' in data:
        log.tags = json.dumps(data['tags'], ensure_ascii=False)
    log.updated_at = dt.utcnow()
    db.session.commit()
    _upsert_chunk('worklog', log.id)
    return jsonify(log.to_dict(include_content=True))


@app.route('/api/worklogs/<int:log_id>', methods=['DELETE'])
def api_worklog_delete(log_id):
    log = db.session.get(WorkLog, log_id) or abort(404)
    _delete_chunk('worklog', log.id)
    db.session.delete(log)
    db.session.commit()
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════════════════════════════
# AI 어시스턴트 (RAG 채팅)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/chat')
def chat_page():
    synced = VectorChunk.query.count()
    return render_template('chat.html', synced_count=synced)


@app.route('/api/chat/sync')
def api_chat_sync():
    """전체 데이터를 임베딩해서 vector_chunk 저장 (SSE 스트림)"""

    task_id = uuid.uuid4().hex
    q = queue.Queue()
    _task_queues[task_id] = q

    def run():
        from rag_service import sync_vectors
        with app.app_context():
            try:
                def cb(current, total, label):
                    q.put({'type': 'progress', 'current': current, 'total': total, 'label': label})
                total = sync_vectors(GEMINI_API_KEY, progress_cb=cb)
                q.put({'type': 'done', 'total': total})
            except Exception as e:
                logger.error("RAG 동기화 오류: %s", e)
                q.put({'type': 'error', 'msg': str(e)})

    threading.Thread(target=run, daemon=True).start()

    def generate():
        yield f"data: {json.dumps({'type': 'started', 'task_id': task_id})}\n\n"
        try:
            while True:
                try:
                    msg = q.get(timeout=120)
                    yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"
                    if msg.get('type') in ('done', 'error'):
                        break
                except queue.Empty:
                    yield 'data: {"type":"keepalive"}\n\n'
        finally:
            _task_queues.pop(task_id, None)

    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'},
    )


@app.route('/api/chat/ask', methods=['POST'])
def api_chat_ask():
    if not GEMINI_API_KEY:
        return jsonify({'error': 'Gemini API 키가 설정되지 않았습니다. GEMINI_API_KEY 환경변수를 설정해주세요.'}), 400
    data    = request.get_json()
    query   = data.get('query', '').strip()
    history = data.get('history', [])
    if not query:
        return jsonify({'error': '질문을 입력해주세요'}), 400

    synced = VectorChunk.query.count()
    if synced == 0:
        return jsonify({'error': '데이터 동기화가 필요합니다. 먼저 동기화를 실행하세요.'}), 400

    from rag_service import answer
    try:
        # PyInstaller 빌드(배포 바이너리)에선 Ollama 없으므로 항상 Gemini 사용
        _use_local = False if getattr(sys, 'frozen', False) else _is_local_request()
        result = answer(query, history, GEMINI_API_KEY, use_local=_use_local)
        return jsonify(result)
    except Exception as e:
        logger.error("RAG 답변 오류: %s", e)
        return jsonify({'error': str(e)}), 500


@app.route('/api/chat/status')
def api_chat_status():
    count = VectorChunk.query.count()
    return jsonify({'synced_count': count})


if __name__ == '__main__':
    if _FROZEN:
        import socket, webbrowser
        port = int(os.environ.get('PORT', 5000))
        host = os.environ.get('HOST', '127.0.0.1')
        def _is_port_in_use(p):
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                return s.connect_ex(('127.0.0.1', p)) == 0
        if _is_port_in_use(port):
            webbrowser.open(f'http://127.0.0.1:{port}')
        else:
            threading.Thread(target=lambda: (
                __import__('time').sleep(1.5),
                webbrowser.open(f'http://127.0.0.1:{port}')
            ), daemon=True).start()
            app.run(debug=False, host=host, port=port, use_reloader=False)
    else:
        port = int(os.environ.get('PORT', 5000))
        host = os.environ.get('HOST', '0.0.0.0')
        app.run(debug=True, host=host, port=port)
